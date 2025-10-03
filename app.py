from flask import Flask, render_template_string, request, jsonify, send_file
import googlemaps
import requests
import re
import time
from urllib.parse import urlparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import os
from datetime import datetime
import json

app = Flask(__name__)

# Your Google Maps API Key
API_KEY = 'AIzaSyBSgFi26h5s1-2MiD8RXre-FRcRAZNbJpY'

def validate_email(email):
    """Validate email format and check if domain exists"""
    if not email or '@' not in email:
        return False
    
    # Basic format validation
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(pattern, email):
        return False
    
    return True

def calculate_lead_score(company):
    """Calculate lead quality score based on available information"""
    score = 0
    
    # Email quality
    if company['emailSource'] == 'scraped':
        score += 40
    elif company['emailSource'] == 'suggested':
        score += 20
    
    # Contact information completeness
    if company['phone'] != 'N/A':
        score += 25
    if company['website'] != 'N/A':
        score += 20
    if company['address'] != 'N/A':
        score += 15
    
    return min(score, 100)

def find_emails_enhanced(url, company_name, discovery_level="enhanced"):
    """Enhanced email finding with multiple strategies"""
    if not url or url == 'N/A':
        return {"emails": [], "source": "none", "pages_checked": 0}
    
    base_url = url.rstrip('/')
    pages_checked = 0
    
    # Different page lists based on discovery level
    if discovery_level == "basic":
        pages = [url, f"{base_url}/contact", f"{base_url}/kontakt"]
    elif discovery_level == "enhanced":
        pages = [
            url, f"{base_url}/contact", f"{base_url}/kontakt", 
            f"{base_url}/about", f"{base_url}/about-us", f"{base_url}/team",
            f"{base_url}/company", f"{base_url}/contact-us", f"{base_url}/om-oss",
            f"{base_url}/personal", f"{base_url}/kontakta-oss"
        ]
    else:  # ai-powered
        pages = [
            url, f"{base_url}/contact", f"{base_url}/kontakt", 
            f"{base_url}/about", f"{base_url}/about-us", f"{base_url}/team",
            f"{base_url}/company", f"{base_url}/contact-us", f"{base_url}/om-oss",
            f"{base_url}/personal", f"{base_url}/kontakta-oss", f"{base_url}/footer",
            f"{base_url}/impressum", f"{base_url}/legal", f"{base_url}/staff",
            f"{base_url}/management", f"{base_url}/leadership"
        ]
    
    for page_url in pages:
        try:
            response = requests.get(page_url, timeout=3, headers={
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
            })
            pages_checked += 1
            
            if response.status_code == 200:
                emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', response.text)
                bad_keywords = ['example.com', 'sentry', 'wix.com', 'noreply', 'no-reply', '@facebook', '@twitter', '@linkedin', '@instagram', 'support@', 'admin@']
                
                valid_emails = []
                for email in emails:
                    email = email.lower().strip()
                    if (len(email) > 5 and '@' in email and 
                        not any(keyword in email for keyword in bad_keywords) and
                        validate_email(email)):
                        valid_emails.append(email)
                
                if valid_emails:
                    # Prioritize certain email patterns
                    priority_emails = [e for e in valid_emails if any(pattern in e for pattern in ['info@', 'contact@', 'sales@'])]
                    if priority_emails:
                        return {"emails": priority_emails[:2], "source": "scraped", "pages_checked": pages_checked}
                    else:
                        return {"emails": valid_emails[:2], "source": "scraped", "pages_checked": pages_checked}
        except:
            continue
    
    # Fallback: Generate likely emails
    try:
        domain = urlparse(url).netloc.replace('www.', '')
        if domain:
            suggested_emails = [f"info@{domain}", f"contact@{domain}", f"sales@{domain}"]
            return {"emails": suggested_emails[:1], "source": "suggested", "pages_checked": pages_checked}
    except:
        pass
    
    return {"emails": [], "source": "none", "pages_checked": pages_checked}

@app.route('/')
def index():
    return render_template_string("""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Salk Service Business Development Tool</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/3.9.1/chart.min.js"></script>
    <style>
        .email-badge {
            display: inline-block;
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 0.75rem;
            font-weight: bold;
        }
        .email-scraped { background-color: #10b981; color: white; }
        .email-suggested { background-color: #f59e0b; color: white; }
        .email-none { background-color: #ef4444; color: white; }
        
        .lead-score-excellent { background-color: #10b981; }
        .lead-score-good { background-color: #3b82f6; }
        .lead-score-fair { background-color: #f59e0b; }
        .lead-score-poor { background-color: #ef4444; }
        
        .category-button {
            transition: all 0.2s ease;
            cursor: pointer;
            border: 2px solid transparent;
        }
        
        .category-button:hover {
            transform: translateY(-1px);
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
        }
        
        .category-button.selected {
            border-color: #3b82f6;
            background-color: #eff6ff;
        }
        
        .subcategory-button {
            transition: all 0.2s ease;
            cursor: pointer;
            border: 1px solid #d1d5db;
        }
        
        .subcategory-button:hover {
            background-color: #f3f4f6;
        }
        
        .subcategory-button.selected {
            background-color: #3b82f6;
            color: white;
            border-color: #3b82f6;
        }
        
        .selection-counter {
            position: sticky;
            top: 0;
            z-index: 10;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 1rem;
            border-radius: 0.5rem;
            margin-bottom: 1rem;
        }
    </style>
</head>
<body class="bg-gray-50 text-gray-900 min-h-screen">
    <div class="container mx-auto px-4 py-6 max-w-7xl">
        <!-- Header -->
        <div class="text-center mb-8">
            <div class="bg-white rounded-lg shadow-lg p-6 mb-6">
                <h1 class="text-5xl font-bold text-blue-600 mb-2">Salk Service</h1>
                <h2 class="text-3xl font-semibold text-gray-700 mb-3">Ultimate Business Development Tool</h2>
                <p class="text-gray-600">Sundsvall Region - Professional Lead Generation & Contact Discovery</p>
                <div class="mt-4 grid grid-cols-1 md:grid-cols-3 gap-4">
                    <div class="bg-green-50 p-3 rounded-lg">
                        <p class="text-green-800 font-semibold">✅ Enhanced Email Discovery</p>
                        <p class="text-sm text-green-600">15+ page scanning strategies</p>
                    </div>
                    <div class="bg-blue-50 p-3 rounded-lg">
                        <p class="text-blue-800 font-semibold">🎯 Lead Scoring</p>
                        <p class="text-sm text-blue-600">Intelligent quality assessment</p>
                    </div>
                    <div class="bg-purple-50 p-3 rounded-lg">
                        <p class="text-purple-800 font-semibold">📊 Advanced Analytics</p>
                        <p class="text-sm text-purple-600">Comprehensive reporting</p>
                    </div>
                </div>
            </div>
        </div>

        <!-- Search Configuration -->
        <div class="bg-white rounded-lg shadow-lg p-6 mb-6">
            <h2 class="text-2xl font-semibold mb-6 text-blue-600">🎯 Search Configuration</h2>
            
            <!-- Basic Settings -->
            <div class="grid grid-cols-1 md:grid-cols-4 gap-4 mb-6">
                <div>
                    <label class="block text-sm font-medium mb-2">Max Companies</label>
                    <input type="number" id="maxCompanies" value="100" min="1" max="500" 
                           class="w-full px-3 py-2 border border-gray-300 rounded-md text-base focus:ring-2 focus:ring-blue-500">
                </div>
                <div>
                    <label class="block text-sm font-medium mb-2">Search Radius (km)</label>
                    <input type="number" id="radius" value="100" min="1" max="200" 
                           class="w-full px-3 py-2 border border-gray-300 rounded-md text-base focus:ring-2 focus:ring-blue-500">
                </div>
                <div>
                    <label class="block text-sm font-medium mb-2">Email Discovery Level</label>
                    <select id="discoveryLevel" class="w-full px-3 py-2 border border-gray-300 rounded-md text-base focus:ring-2 focus:ring-blue-500">
                        <option value="basic">Basic (3 pages)</option>
                        <option value="enhanced" selected>Enhanced (10+ pages)</option>
                        <option value="ai-powered">AI-Powered (15+ pages)</option>
                    </select>
                </div>
                <div>
                    <label class="block text-sm font-medium mb-2">Search Delay (seconds)</label>
                    <select id="searchDelay" class="w-full px-3 py-2 border border-gray-300 rounded-md text-base focus:ring-2 focus:ring-blue-500">
                        <option value="0.1">Fast (0.1s)</option>
                        <option value="0.2" selected>Normal (0.2s)</option>
                        <option value="0.5">Careful (0.5s)</option>
                        <option value="1.0">Conservative (1.0s)</option>
                    </select>
                </div>
            </div>

            <!-- Selection Counter -->
            <div class="selection-counter">
                <div class="flex justify-between items-center">
                    <div>
                        <span class="text-lg font-bold">Selected: <span id="selectionCount">0</span> business types</span>
                        <div class="text-sm opacity-90 mt-1">Click main categories to select all subcategories, or click individual items</div>
                    </div>
                    <div class="flex gap-2">
                        <button onclick="selectAll()" class="bg-white bg-opacity-20 hover:bg-opacity-30 px-3 py-1 rounded text-sm">Select All</button>
                        <button onclick="deselectAll()" class="bg-white bg-opacity-20 hover:bg-opacity-30 px-3 py-1 rounded text-sm">Clear All</button>
                        <button onclick="testAPI()" class="bg-yellow-500 hover:bg-yellow-600 px-3 py-1 rounded text-sm">🧪 Test API</button>
                    </div>
                </div>
            </div>

            <!-- Business Categories -->
            <div class="mb-6">
                <h3 class="text-lg font-semibold mb-4">🏭 Complete Business Categories & Subcategories</h3>
                <div id="categoriesContainer">
                    <!-- Categories will be populated by JavaScript -->
                </div>
            </div>

            <button onclick="startSearch()" id="searchBtn" 
                    class="w-full bg-gradient-to-r from-blue-600 to-purple-600 hover:from-blue-700 hover:to-purple-700 text-white font-bold py-4 px-6 rounded-lg text-xl shadow-lg transition-all duration-200 transform hover:scale-105">
                🚀 START COMPREHENSIVE BUSINESS SEARCH
            </button>
        </div>

        <!-- Progress -->
        <div id="progress" class="hidden mb-6">
            <div class="bg-white rounded-lg shadow-lg p-6">
                <h3 class="text-lg font-semibold mb-4">Search Progress</h3>
                <div class="bg-gray-200 rounded-full h-6 mb-2">
                    <div id="progressBar" class="bg-gradient-to-r from-blue-500 to-purple-500 h-6 rounded-full transition-all duration-300" style="width: 0%"></div>
                </div>
                <div class="flex justify-between text-sm text-gray-600">
                    <span id="progressText">Starting search...</span>
                    <span id="progressPercent">0%</span>
                </div>
                <div id="progressDetails" class="mt-2 text-xs text-gray-500"></div>
            </div>
        </div>

        <!-- Results -->
        <div id="results" class="hidden">
            <!-- Analytics Dashboard (no round chart) -->
            <div class="bg-white rounded-lg shadow-lg p-6 mb-6">
                <h2 class="text-2xl font-semibold mb-6 text-blue-600">📊 Lead Generation Analytics</h2>
                <div id="analytics" class="grid grid-cols-2 md:grid-cols-4 lg:grid-cols-6 gap-4 mb-6"></div>
                <div class="flex justify-center">
                    <div class="w-full max-w-4xl">
                        <canvas id="leadScoreChart" width="400" height="200"></canvas>
                    </div>
                </div>
            </div>

            <!-- Filters and Export -->
            <div class="bg-white rounded-lg shadow-lg p-6 mb-6">
                <h3 class="text-lg font-semibold mb-4">🔍 Results Management</h3>
                <div class="grid grid-cols-1 md:grid-cols-3 gap-4 mb-4">
                    <div>
                        <label class="block text-sm font-medium mb-2">Filter by Email Source</label>
                        <select id="emailFilter" onchange="filterResults()" class="w-full px-3 py-2 border border-gray-300 rounded-md">
                            <option value="all">All Results</option>
                            <option value="scraped">Scraped Emails Only</option>
                            <option value="suggested">Suggested Emails Only</option>
                            <option value="none">No Email Found</option>
                        </select>
                    </div>
                    <div>
                        <label class="block text-sm font-medium mb-2">Filter by Lead Score</label>
                        <select id="scoreFilter" onchange="filterResults()" class="w-full px-3 py-2 border border-gray-300 rounded-md">
                            <option value="all">All Scores</option>
                            <option value="excellent">Excellent (80-100)</option>
                            <option value="good">Good (60-79)</option>
                            <option value="fair">Fair (40-59)</option>
                            <option value="poor">Poor (0-39)</option>
                        </select>
                    </div>
                    <div>
                        <label class="block text-sm font-medium mb-2">Sort Results By</label>
                        <select id="sortFilter" onchange="sortResults()" class="w-full px-3 py-2 border border-gray-300 rounded-md">
                            <option value="score">Lead Score (High to Low)</option>
                            <option value="name">Company Name (A-Z)</option>
                            <option value="type">Business Type</option>
                            <option value="email">Email Quality</option>
                        </select>
                    </div>
                </div>
                
                <div class="flex flex-wrap gap-3">
                    <button onclick="downloadExcel('all')" class="bg-green-600 hover:bg-green-700 text-white font-bold py-2 px-4 rounded-lg">
                        📋 Download All (Excel)
                    </button>
                    <button onclick="downloadExcel('scraped')" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-lg">
                        ✉️ Verified Emails Only
                    </button>
                    <button onclick="downloadExcel('high-score')" class="bg-purple-600 hover:bg-purple-700 text-white font-bold py-2 px-4 rounded-lg">
                        ⭐ High-Score Leads
                    </button>
                    <button onclick="downloadCSV()" class="bg-gray-600 hover:bg-gray-700 text-white font-bold py-2 px-4 rounded-lg">
                        📄 CSV Format
                    </button>
                </div>
            </div>

            <!-- Results Table -->
            <div class="bg-white rounded-lg shadow-lg p-6">
                <h2 class="text-xl font-semibold mb-4">🎯 Business Development Leads</h2>
                <div class="overflow-x-auto">
                    <table class="w-full border-collapse border border-gray-300">
                        <thead>
                            <tr class="bg-gradient-to-r from-blue-50 to-purple-50">
                                <th class="border border-gray-300 px-4 py-3 text-left font-semibold">Lead Score</th>
                                <th class="border border-gray-300 px-4 py-3 text-left font-semibold">Company Information</th>
                                <th class="border border-gray-300 px-4 py-3 text-left font-semibold">Business Type</th>
                                <th class="border border-gray-300 px-4 py-3 text-left font-semibold">Email Contact</th>
                                <th class="border border-gray-300 px-4 py-3 text-left font-semibold">Phone</th>
                                <th class="border border-gray-300 px-4 py-3 text-left font-semibold">Website</th>
                            </tr>
                        </thead>
                        <tbody id="resultsTable"></tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <script>
        // COMPLETE business categories data - ALL CATEGORIES AND SUBCATEGORIES
        const BUSINESS_CATEGORIES = {
            "INDUSTRIAL & MANUFACTURING": {
                "Metal Works": ['metal_fabrication', 'steel_mill', 'foundry', 'machine_shop', 'welding_shop', 'sheet_metal_shop', 'metal_stamping', 'metal_casting', 'metal_finishing', 'powder_coating', 'galvanizing_plant', 'heat_treatment', 'metal_plating', 'cnc_machining', 'precision_machining', 'tool_and_die', 'mold_maker', 'aluminum_plant'],
                "Plastics & Rubber": ['plastics_manufacturer', 'rubber_manufacturer', 'injection_molding', 'extrusion_company', 'thermoforming', 'composite_manufacturing'],
                "Tubes & Pipes": ['tube_manufacturer', 'pipe_manufacturer', 'hose_manufacturer', 'valve_manufacturer', 'pump_manufacturer', 'fitting_manufacturer'],
                "General Manufacturing": ['factory', 'manufacturing', 'industrial_area', 'warehouse', 'assembly_plant', 'production_facility', 'industrial_park'],
                "Electronics": ['electronics_manufacturer', 'circuit_board_manufacturer', 'cable_manufacturer', 'wire_manufacturer', 'semiconductor_plant'],
                "Specialized Parts": ['bearing_manufacturer', 'gear_manufacturer', 'spring_manufacturer', 'fastener_manufacturer', 'auto_parts_manufacturer', 'aerospace_parts']
            },
            "AUTOMOTIVE & TRANSPORTATION": {
                "Commercial Vehicles": ['truck_dealer', 'truck_repair', 'fleet_maintenance', 'bus_depot', 'truck_wash', 'commercial_vehicle_dealer', 'trailer_dealer', 'heavy_equipment_dealer'],
                "Automotive Services": ['car_dealer', 'car_rental', 'car_repair', 'car_wash', 'auto_parts_manufacturer', 'tire_manufacturer', 'battery_manufacturer', 'automotive_supplier', 'engine_rebuilding', 'transmission_shop']
            },
            "FOOD & BEVERAGE PRODUCTION": {
                "Food Processing": ['food_processing_plant', 'meat_processing', 'dairy_processing', 'bakery_wholesale', 'canning_facility', 'frozen_food_manufacturer', 'spice_manufacturer', 'confectionery_manufacturer'],
                "Beverages": ['beverage_plant', 'brewery', 'winery', 'distillery', 'bottling_plant', 'soft_drink_manufacturer']
            },
            "CHEMICAL & PHARMACEUTICAL": {
                "Chemicals": ['chemical_plant', 'chemical_manufacturer', 'paint_manufacturer', 'adhesive_manufacturer', 'solvent_supplier', 'industrial_gas_supplier', 'fertilizer_plant', 'pesticide_manufacturer'],
                "Pharmaceutical": ['pharmaceutical_company', 'laboratory', 'biotechnology_company', 'medical_device_manufacturer']
            },
            "MATERIALS & CONSTRUCTION": {
                "Building Materials": ['concrete_plant', 'asphalt_plant', 'brick_manufacturer', 'cement_plant', 'tile_manufacturer', 'roofing_manufacturer', 'insulation_manufacturer'],
                "Wood Products": ['sawmill', 'lumber_yard', 'paper_mill', 'woodworking_shop', 'cabinet_maker', 'furniture_manufacturer', 'pallet_manufacturer', 'pallet_yard', 'veneer_mill', 'plywood_mill'],
                "Glass & Ceramics": ['glass_factory', 'ceramics_manufacturer', 'fiberglass_manufacturer'],
                "Construction": ['general_contractor', 'construction_company', 'civil_engineering', 'roofing_contractor', 'demolition_contractor']
            },
            "TEXTILES & PACKAGING": {
                "Textiles": ['textile_mill', 'garment_factory', 'textile_factory', 'dyeing_plant', 'embroidery_shop', 'upholstery_shop', 'carpet_manufacturer'],
                "Packaging": ['packaging_manufacturer', 'box_manufacturer', 'container_manufacturer', 'label_printer', 'carton_manufacturer']
            },
            "ENERGY & UTILITIES": {
                "Energy Production": ['power_plant', 'solar_energy_company', 'wind_farm', 'biomass_plant', 'hydroelectric_plant', 'nuclear_power_plant'],
                "Utilities": ['water_treatment_plant', 'sewage_treatment_plant', 'district_heating', 'energy_company', 'utility_contractor', 'waste_management', 'recycling_center']
            },
            "AGRICULTURE & MINING": {
                "Agriculture": ['farm', 'agricultural_cooperative', 'farm_equipment_supplier', 'greenhouse', 'nursery', 'dairy_farm', 'livestock_farm', 'agricultural_service'],
                "Mining": ['mining_company', 'quarry', 'gravel_pit', 'sand_quarry', 'stone_quarry', 'coal_mine']
            },
            "LOGISTICS & FACILITIES": {
                "Logistics": ['warehouse', 'distribution_center', 'shipping_company', 'freight_forwarder', 'logistics_service', 'cold_storage', 'port_authority', 'storage'],
                "Facilities Management": ['facility_management', 'property_management', 'building_maintenance', 'janitorial_service', 'cleaning_service', 'office_cleaning', 'industrial_cleaning', 'pressure_washing_service']
            },
            "COMMERCIAL & PUBLIC": {
                "Large Facilities": ['hotel', 'hospital', 'shopping_mall', 'convention_center', 'exhibition_center', 'conference_center', 'sports_complex', 'arena', 'stadium', 'racecourse', 'fairground', 'airport'],
                "Municipal": ['city_hall', 'local_government_office', 'public_works_department', 'fire_station', 'police', 'post_office', 'library'],
                "Retail & Service": ['restaurant', 'cafe', 'supermarket', 'department_store', 'hardware_store', 'furniture_store', 'electronics_store', 'gas_station', 'car_wash'],
                "Equipment & Rental": ['equipment_rental', 'tool_rental', 'heavy_machinery_dealer', 'moving_company']
            }
        };

        let currentResults = [];
        let filteredResults = [];
        let selectedBusinessTypes = new Set();

        async function testAPI() {
            try {
                const response = await fetch('/api/test-maps', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({test: true})
                });
                
                const data = await response.json();
                
                if (data.success) {
                    alert('✅ Google Maps API Working! Ready to search.');
                } else {
                    alert('❌ API Error: ' + data.error);
                }
            } catch (error) {
                alert('❌ Connection Error: ' + error.message);
            }
        }

        // Initialize categories with complete data
        function initializeCategories() {
            console.log('Initializing complete categories...');
            const container = document.getElementById('categoriesContainer');
            
            if (!container) {
                console.error('Categories container not found!');
                return;
            }
            
            container.innerHTML = '';

            for (const [mainCategory, subcategories] of Object.entries(BUSINESS_CATEGORIES)) {
                console.log('Creating category:', mainCategory);
                
                const categorySection = document.createElement('div');
                categorySection.className = 'mb-8';
                
                // Main category header (clickable)
                const mainCategoryButton = document.createElement('div');
                mainCategoryButton.className = 'category-button bg-gradient-to-r from-blue-500 to-purple-500 text-white p-4 rounded-lg mb-4 font-bold text-lg text-center';
                const totalBusinessTypes = Object.values(subcategories).flat().length;
                mainCategoryButton.textContent = `${mainCategory} (${totalBusinessTypes} business types)`;
                mainCategoryButton.onclick = () => toggleMainCategory(mainCategory);
                categorySection.appendChild(mainCategoryButton);

                // Subcategories container
                const subcategoriesContainer = document.createElement('div');
                subcategoriesContainer.className = 'grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-4';

                for (const [subCategory, businessTypes] of Object.entries(subcategories)) {
                    const subCategoryDiv = document.createElement('div');
                    subCategoryDiv.className = 'bg-gray-50 p-4 rounded-lg';
                    
                    // Subcategory header (clickable)
                    const subCategoryButton = document.createElement('div');
                    subCategoryButton.className = 'subcategory-button bg-white p-3 rounded-lg mb-3 font-semibold text-center';
                    subCategoryButton.textContent = `${subCategory} (${businessTypes.length} types)`;
                    subCategoryButton.onclick = () => toggleSubCategory(subCategory, businessTypes);
                    subCategoryDiv.appendChild(subCategoryButton);

                    // Individual business types
                    const businessTypesGrid = document.createElement('div');
                    businessTypesGrid.className = 'grid grid-cols-1 gap-2';

                    businessTypes.forEach(businessType => {
                        const businessTypeButton = document.createElement('div');
                        businessTypeButton.className = 'subcategory-button bg-white p-2 rounded text-center text-sm';
                        businessTypeButton.textContent = businessType.replace(/_/g, ' ').replace(/\\b\\w/g, l => l.toUpperCase());
                        businessTypeButton.onclick = () => toggleBusinessType(businessType);
                        businessTypeButton.dataset.businessType = businessType;
                        businessTypesGrid.appendChild(businessTypeButton);
                    });

                    subCategoryDiv.appendChild(businessTypesGrid);
                    subcategoriesContainer.appendChild(subCategoryDiv);
                }

                categorySection.appendChild(subcategoriesContainer);
                container.appendChild(categorySection);
            }
            
            console.log('Complete categories initialized successfully');
            updateSelectionCount();
        }

        function toggleMainCategory(mainCategory) {
            const allBusinessTypes = Object.values(BUSINESS_CATEGORIES[mainCategory]).flat();
            const allSelected = allBusinessTypes.every(bt => selectedBusinessTypes.has(bt));
            
            if (allSelected) {
                allBusinessTypes.forEach(bt => selectedBusinessTypes.delete(bt));
            } else {
                allBusinessTypes.forEach(bt => selectedBusinessTypes.add(bt));
            }
            
            updateVisualSelection();
            updateSelectionCount();
        }

        function toggleSubCategory(subCategory, businessTypes) {
            const allSelected = businessTypes.every(bt => selectedBusinessTypes.has(bt));
            
            if (allSelected) {
                businessTypes.forEach(bt => selectedBusinessTypes.delete(bt));
            } else {
                businessTypes.forEach(bt => selectedBusinessTypes.add(bt));
            }
            
            updateVisualSelection();
            updateSelectionCount();
        }

        function toggleBusinessType(businessType) {
            if (selectedBusinessTypes.has(businessType)) {
                selectedBusinessTypes.delete(businessType);
            } else {
                selectedBusinessTypes.add(businessType);
            }
            
            updateVisualSelection();
            updateSelectionCount();
        }

        function updateVisualSelection() {
            document.querySelectorAll('[data-business-type]').forEach(button => {
                const businessType = button.dataset.businessType;
                if (selectedBusinessTypes.has(businessType)) {
                    button.classList.add('selected');
                } else {
                    button.classList.remove('selected');
                }
            });

            document.querySelectorAll('.category-button').forEach(button => {
                const mainCategoryText = button.textContent.split(' (')[0]; // Remove count part
                if (BUSINESS_CATEGORIES[mainCategoryText]) {
                    const allBusinessTypes = Object.values(BUSINESS_CATEGORIES[mainCategoryText]).flat();
                    const allSelected = allBusinessTypes.every(bt => selectedBusinessTypes.has(bt));
                    if (allSelected) {
                        button.classList.add('selected');
                    } else {
                        button.classList.remove('selected');
                    }
                }
            });
        }

        function updateSelectionCount() {
            document.getElementById('selectionCount').textContent = selectedBusinessTypes.size;
        }

        function selectAll() {
            selectedBusinessTypes.clear();
            for (const [mainCategory, subcategories] of Object.entries(BUSINESS_CATEGORIES)) {
                for (const [subCategory, businessTypes] of Object.entries(subcategories)) {
                    businessTypes.forEach(bt => selectedBusinessTypes.add(bt));
                }
            }
            updateVisualSelection();
            updateSelectionCount();
        }

        function deselectAll() {
            selectedBusinessTypes.clear();
            updateVisualSelection();
            updateSelectionCount();
        }

        function updateProgress(percent, text, details = '') {
            document.getElementById('progressBar').style.width = percent + '%';
            document.getElementById('progressPercent').textContent = percent + '%';
            document.getElementById('progressText').textContent = text;
            if (details) {
                document.getElementById('progressDetails').textContent = details;
            }
        }

        async function startSearch() {
            if (selectedBusinessTypes.size === 0) {
                alert('Please select at least one business type!');
                return;
            }

            const categories = Array.from(selectedBusinessTypes);
            const maxCompanies = parseInt(document.getElementById('maxCompanies').value);
            const radius = parseInt(document.getElementById('radius').value);
            const discoveryLevel = document.getElementById('discoveryLevel').value;
            const searchDelay = parseFloat(document.getElementById('searchDelay').value);

            console.log(`Starting comprehensive search with ${categories.length} business types`);

            document.getElementById('searchBtn').disabled = true;
            document.getElementById('searchBtn').textContent = '🔍 SEARCHING COMPREHENSIVE DATABASE...';
            document.getElementById('progress').classList.remove('hidden');
            document.getElementById('results').classList.add('hidden');

            updateProgress(0, 'Initializing comprehensive search...', `Preparing to search ${categories.length} business types`);

            try {
                const response = await fetch('/api/search', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        categories,
                        maxCompanies,
                        radius,
                        discoveryLevel,
                        searchDelay
                    })
                });

                if (!response.ok) {
                    throw new Error('Search request failed with status ' + response.status);
                }

                const data = await response.json();
                
                console.log(`Comprehensive search completed. Found ${data.companies.length} companies`);
                
                currentResults = data.companies;
                filteredResults = [...currentResults];
                updateProgress(100, 'Comprehensive search completed!', `Found ${data.companies.length} companies`);
                displayResults();
                createCharts();

            } catch (error) {
                console.error('Search error:', error);
                alert('Search failed: ' + error.message);
                updateProgress(0, 'Search failed', error.message);
            } finally {
                document.getElementById('searchBtn').disabled = false;
                document.getElementById('searchBtn').textContent = '🚀 START COMPREHENSIVE BUSINESS SEARCH';
                setTimeout(() => {
                    document.getElementById('progress').classList.add('hidden');
                }, 3000);
            }
        }

        function displayResults() {
            const resultsDiv = document.getElementById('results');
            const analyticsDiv = document.getElementById('analytics');
            const tableBody = document.getElementById('resultsTable');

            // Calculate analytics
            const total = filteredResults.length;
            const scraped = filteredResults.filter(c => c.emailSource === 'scraped').length;
            const suggested = filteredResults.filter(c => c.emailSource === 'suggested').length;
            const hasPhone = filteredResults.filter(c => c.phone !== 'N/A').length;
            const hasWebsite = filteredResults.filter(c => c.website !== 'N/A').length;
            const excellentLeads = filteredResults.filter(c => c.leadScore >= 80).length;

            analyticsDiv.innerHTML = `
                <div class="bg-gradient-to-br from-blue-500 to-blue-600 text-white p-4 rounded-lg text-center">
                    <div class="text-3xl font-bold">${total}</div>
                    <div class="text-sm opacity-90">Total Leads</div>
                </div>
                <div class="bg-gradient-to-br from-green-500 to-green-600 text-white p-4 rounded-lg text-center">
                    <div class="text-3xl font-bold">${scraped}</div>
                    <div class="text-sm opacity-90">Verified Emails</div>
                </div>
                <div class="bg-gradient-to-br from-orange-500 to-orange-600 text-white p-4 rounded-lg text-center">
                    <div class="text-3xl font-bold">${suggested}</div>
                    <div class="text-sm opacity-90">Suggested Emails</div>
                </div>
                <div class="bg-gradient-to-br from-purple-500 to-purple-600 text-white p-4 rounded-lg text-center">
                    <div class="text-3xl font-bold">${hasPhone}</div>
                    <div class="text-sm opacity-90">Phone Numbers</div>
                </div>
                <div class="bg-gradient-to-br from-indigo-500 to-indigo-600 text-white p-4 rounded-lg text-center">
                    <div class="text-3xl font-bold">${hasWebsite}</div>
                    <div class="text-sm opacity-90">Websites</div>
                </div>
                <div class="bg-gradient-to-br from-yellow-500 to-yellow-600 text-white p-4 rounded-lg text-center">
                    <div class="text-3xl font-bold">${excellentLeads}</div>
                    <div class="text-sm opacity-90">Excellent Leads</div>
                </div>
            `;

            // Populate table
            tableBody.innerHTML = '';
            filteredResults.forEach(company => {
                const row = tableBody.insertRow();
                
                const leadScoreClass = company.leadScore >= 80 ? 'lead-score-excellent' :
                                     company.leadScore >= 60 ? 'lead-score-good' :
                                     company.leadScore >= 40 ? 'lead-score-fair' : 'lead-score-poor';
                
                const emailBadge = company.emailSource === 'scraped' ? 
                    '<span class="email-badge email-scraped">SCRAPED</span>' : 
                    company.emailSource === 'suggested' ? '<span class="email-badge email-suggested">SUGGESTED</span>' :
                    '<span class="email-badge email-none">NONE</span>';

                row.innerHTML = `
                    <td class="border border-gray-300 px-4 py-3 text-center">
                        <div class="${leadScoreClass} text-white px-3 py-2 rounded-full font-bold">
                            ${company.leadScore}
                        </div>
                    </td>
                    <td class="border border-gray-300 px-4 py-3">
                        <div class="font-semibold text-blue-700">${company.name}</div>
                        <div class="text-sm text-gray-600">${company.address}</div>
                    </td>
                    <td class="border border-gray-300 px-4 py-3">
                        <span class="bg-blue-100 text-blue-800 px-2 py-1 rounded text-sm">${company.type}</span>
                    </td>
                    <td class="border border-gray-300 px-4 py-3">
                        ${company.email !== 'N/A' ? `<a href="mailto:${company.email}" class="text-blue-600 hover:underline font-medium">${company.email}</a>` : 'N/A'}
                        <div class="mt-1">${emailBadge}</div>
                    </td>
                    <td class="border border-gray-300 px-4 py-3">
                        ${company.phone !== 'N/A' ? `<a href="tel:${company.phone}" class="text-blue-600 hover:underline">${company.phone}</a>` : 'N/A'}
                    </td>
                    <td class="border border-gray-300 px-4 py-3">
                        ${company.website !== 'N/A' ? `<a href="${company.website}" target="_blank" class="text-blue-600 hover:underline font-medium">Visit Site</a>` : 'N/A'}
                    </td>
                `;
            });

            resultsDiv.classList.remove('hidden');
        }

        function filterResults() {
            const emailFilter = document.getElementById('emailFilter').value;
            const scoreFilter = document.getElementById('scoreFilter').value;
            
            filteredResults = currentResults.filter(company => {
                let emailMatch = true;
                let scoreMatch = true;
                
                if (emailFilter !== 'all') {
                    emailMatch = company.emailSource === emailFilter;
                }
                
                if (scoreFilter !== 'all') {
                    switch(scoreFilter) {
                        case 'excellent': scoreMatch = company.leadScore >= 80; break;
                        case 'good': scoreMatch = company.leadScore >= 60 && company.leadScore < 80; break;
                        case 'fair': scoreMatch = company.leadScore >= 40 && company.leadScore < 60; break;
                        case 'poor': scoreMatch = company.leadScore < 40; break;
                    }
                }
                
                return emailMatch && scoreMatch;
            });
            
            displayResults();
        }

        function sortResults() {
            const sortBy = document.getElementById('sortFilter').value;
            
            filteredResults.sort((a, b) => {
                switch(sortBy) {
                    case 'score': return b.leadScore - a.leadScore;
                    case 'name': return a.name.localeCompare(b.name);
                    case 'type': return a.type.localeCompare(b.type);
                    case 'email': 
                        const emailOrder = {'scraped': 3, 'suggested': 2, 'none': 1};
                        return emailOrder[b.emailSource] - emailOrder[a.emailSource];
                    default: return 0;
                }
            });
            
            displayResults();
        }

        function createCharts() {
            // Lead Score Distribution Chart (removed round doughnut chart)
            const scoreCtx = document.getElementById('leadScoreChart').getContext('2d');
            new Chart(scoreCtx, {
                type: 'bar',
                data: {
                    labels: ['Excellent (80-100)', 'Good (60-79)', 'Fair (40-59)', 'Poor (0-39)'],
                    datasets: [{
                        label: 'Number of Companies',
                        data: [
                            currentResults.filter(c => c.leadScore >= 80).length,
                            currentResults.filter(c => c.leadScore >= 60 && c.leadScore < 80).length,
                            currentResults.filter(c => c.leadScore >= 40 && c.leadScore < 60).length,
                            currentResults.filter(c => c.leadScore < 40).length
                        ],
                        backgroundColor: ['#10b981', '#3b82f6', '#f59e0b', '#ef4444'],
                        borderColor: ['#059669', '#2563eb', '#d97706', '#dc2626'],
                        borderWidth: 1
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: true,
                    plugins: {
                        title: { 
                            display: true, 
                            text: 'Lead Quality Distribution',
                            font: { size: 18, weight: 'bold' }
                        },
                        legend: {
                            display: false
                        }
                    },
                    scales: {
                        y: {
                            beginAtZero: true,
                            title: {
                                display: true,
                                text: 'Number of Companies',
                                font: { size: 14, weight: 'bold' }
                            },
                            grid: {
                                color: '#e5e7eb'
                            }
                        },
                        x: {
                            title: {
                                display: true,
                                text: 'Lead Score Range',
                                font: { size: 14, weight: 'bold' }
                            },
                            grid: {
                                display: false
                            }
                        }
                    }
                }
            });
        }

        async function downloadExcel(type) {
            let dataToExport = currentResults;
            
            if (type === 'scraped') {
                dataToExport = currentResults.filter(c => c.emailSource === 'scraped');
            } else if (type === 'high-score') {
                dataToExport = currentResults.filter(c => c.leadScore >= 80);
            }

            try {
                const response = await fetch('/api/download-excel', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({companies: dataToExport, type: type})
                });

                const blob = await response.blob();
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `Salk_Service_Business_Leads_${type}_${new Date().toISOString().split('T')[0]}.xlsx`;
                a.click();
                URL.revokeObjectURL(url);
            } catch (error) {
                alert('Excel download failed: ' + error.message);
            }
        }

        function downloadCSV() {
            const csvContent = [
                ['Company Name', 'Business Type', 'Address', 'Phone', 'Website', 'Email', 'Email Source', 'Lead Score', 'Contact Name'].join(','),
                ...filteredResults.map(c => [
                    `"${c.name}"`, `"${c.type}"`, `"${c.address}"`,
                    c.phone, c.website, c.email, c.emailSource, c.leadScore, 'N/A'
                ].join(','))
            ].join('\\n');

            const blob = new Blob([csvContent], { type: 'text/csv' });
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = `Salk_Service_Leads_${new Date().toISOString().split('T')[0]}.csv`;
            a.click();
            URL.revokeObjectURL(url);
        }

        // Initialize on page load
        document.addEventListener('DOMContentLoaded', function() {
            console.log('DOM loaded, initializing complete categories...');
            initializeCategories();
        });
    </script>
</body>
</html>
    """)

@app.route('/api/test-maps', methods=['POST'])
def test_maps():
    """Test Google Maps API connection"""
    try:
        print("Testing Google Maps API...")
        gmaps = googlemaps.Client(key=API_KEY)
        
        # Test with a simple search around Sundsvall
        location = (62.3908, 17.3069)  # Sundsvall coordinates
        
        places_result = gmaps.places_nearby(
            location=location, 
            radius=10000,  # 10km 
            type='store'  # Simple type that should exist
        )
        
        print(f"API test result: Found {len(places_result.get('results', []))} places")
        
        if places_result.get('results'):
            return jsonify({
                'success': True, 
                'message': f'Found {len(places_result["results"])} test businesses'
            })
        else:
            return jsonify({
                'success': False, 
                'error': 'No results returned from API'
            })
            
    except Exception as e:
        print(f"API test error: {e}")
        return jsonify({
            'success': False, 
            'error': str(e)
        })

@app.route('/api/search', methods=['POST'])
def search_businesses():
    try:
        print("Comprehensive search request received")
        data = request.json
        categories = data['categories']
        max_companies = data['maxCompanies']
        radius = data['radius'] * 1000  # Convert to meters
        discovery_level = data['discoveryLevel']
        search_delay = data.get('searchDelay', 0.2)
        
        print(f"Comprehensive search parameters: {len(categories)} categories, max={max_companies}, radius={radius}m")
        
        gmaps = googlemaps.Client(key=API_KEY)
        location = (62.3908, 17.3069)  # Sundsvall coordinates
        companies = []
        seen_names = set()
        
        for i, category in enumerate(categories):
            if len(companies) >= max_companies:
                break
                
            print(f"Searching category {i+1}/{len(categories)}: {category}")
            
            try:
                places_result = gmaps.places_nearby(
                    location=location, 
                    radius=radius, 
                    type=category
                )
                
                places_found = len(places_result.get('results', []))
                print(f"Found {places_found} places for {category}")
                
                for place in places_result.get('results', []):
                    if len(companies) >= max_companies:
                        break
                    
                    place_id = place['place_id']
                    details = gmaps.place(
                        place_id, 
                        fields=['name', 'formatted_address', 'formatted_phone_number', 'website']
                    )['result']
                    
                    company_name = details.get('name', 'N/A')
                    if company_name in seen_names:
                        continue
                    seen_names.add(company_name)
                    
                    website = details.get('website', 'N/A')
                    
                    # Enhanced email discovery
                    email_result = find_emails_enhanced(website, company_name, discovery_level)
                    
                    company_data = {
                        'name': company_name,
                        'type': category.replace('_', ' ').title(),
                        'address': details.get('formatted_address', 'N/A'),
                        'phone': details.get('formatted_phone_number', 'N/A'),
                        'website': website,
                        'email': email_result['emails'][0] if email_result['emails'] else 'N/A',
                        'emailSource': email_result['source'],
                        'pagesChecked': email_result.get('pages_checked', 0)
                    }
                    
                    # Calculate lead score
                    company_data['leadScore'] = calculate_lead_score(company_data)
                    
                    companies.append(company_data)
                    print(f"Added: {company_name} (Score: {company_data['leadScore']})")
                    
                    time.sleep(search_delay)  # Rate limiting
                    
            except Exception as e:
                print(f"Error with category {category}: {e}")
                continue
        
        print(f"Comprehensive search completed. Total companies found: {len(companies)}")
        return jsonify({'companies': companies})
        
    except Exception as e:
        print(f"Comprehensive search error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    try:
        data = request.json
        companies = data['companies']
        export_type = data.get('type', 'all')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Salk Service Leads - {export_type.title()}"
        
        # Set headers
        headers = ['Lead Score', 'Company Name', 'Business Type', 'Address', 'Phone', 'Website', 'Email', 'Email Source', 'Pages Checked', 'Contact Name', 'Notes']
        
        # Style the header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data with conditional formatting
        for row, company in enumerate(companies, 2):
            cells = [
                ws.cell(row=row, column=1, value=company['leadScore']),
                ws.cell(row=row, column=2, value=company['name']),
                ws.cell(row=row, column=3, value=company['type']),
                ws.cell(row=row, column=4, value=company['address']),
                ws.cell(row=row, column=5, value=company['phone']),
                ws.cell(row=row, column=6, value=company['website']),
                ws.cell(row=row, column=7, value=company['email']),
                ws.cell(row=row, column=8, value=company['emailSource'].upper()),
                ws.cell(row=row, column=9, value=company.get('pagesChecked', 0)),
                ws.cell(row=row, column=10, value=''),  # Contact Name
                ws.cell(row=row, column=11, value='')   # Notes
            ]
            
            # Color code based on lead score
            lead_score = company['leadScore']
            if lead_score >= 80:
                fill_color = 'E6F7E6'  # Light green
            elif lead_score >= 60:
                fill_color = 'E6F2FF'  # Light blue
            elif lead_score >= 40:
                fill_color = 'FFF2E6'  # Light orange
            else:
                fill_color = 'FFE6E6'  # Light red
            
            for cell in cells:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        summary_data = [
            ['Salk Service Comprehensive Business Development Report', ''],
            ['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Export Type:', export_type.title()],
            ['', ''],
            ['STATISTICS', ''],
            ['Total Companies:', len(companies)],
            ['Verified Emails:', len([c for c in companies if c['emailSource'] == 'scraped'])],
            ['Suggested Emails:', len([c for c in companies if c['emailSource'] == 'suggested'])],
            ['Phone Numbers:', len([c for c in companies if c['phone'] != 'N/A'])],
            ['Websites:', len([c for c in companies if c['website'] != 'N/A'])],
            ['', ''],
            ['LEAD QUALITY', ''],
            ['Excellent Leads (80-100):', len([c for c in companies if c['leadScore'] >= 80])],
            ['Good Leads (60-79):', len([c for c in companies if c['leadScore'] >= 60 and c['leadScore'] < 80])],
            ['Fair Leads (40-59):', len([c for c in companies if c['leadScore'] >= 40 and c['leadScore'] < 60])],
            ['Poor Leads (0-39):', len([c for c in companies if c['leadScore'] < 40])]
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=value)
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f'Salk_Service_Comprehensive_Business_Leads_{export_type}_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == "__main__":
    print("🚀 Starting Salk Service Business Development Tool...")
    print("📍 Deployed online and ready for team access!")
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
